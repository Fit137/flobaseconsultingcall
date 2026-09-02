#!/usr/bin/env python3
"""Convert a GTM dashboard markdown file into a formatted XLSX workbook.

Each level-2 heading (## ...) becomes a worksheet. Markdown tables under that
heading are written as cell ranges; prose and bullets are written as notes rows.

Usage: md_to_xlsx.py <input.md> <output.xlsx> [SheetName=Custom ...]
"""
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
NOTE_FONT = Font(italic=True, size=10, color="404040")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def split_row(line):
    parts = line.strip().strip("|").split("|")
    return [p.strip() for p in parts]


def is_divider(line):
    return bool(re.fullmatch(r"\|[\s:\-|]+\|", line.strip()))


def parse(md_text):
    """Return [(section_title, [block, ...])] where a block is a table or prose."""
    sections = []
    current = ("Overview", [])
    lines = md_text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith("## "):
            if current[1]:
                sections.append(current)
            current = (line[3:].strip(), [])
            i += 1
            continue
        if line.startswith("### "):
            current[1].append(("subhead", line[4:].strip()))
            i += 1
            continue
        if line.strip().startswith("|") and i + 1 < len(lines) and is_divider(lines[i + 1]):
            header = split_row(line)
            rows = []
            i += 2
            while i < len(lines) and lines[i].strip().startswith("|"):
                rows.append(split_row(lines[i]))
                i += 1
            current[1].append(("table", (header, rows)))
            continue
        text = line.strip()
        if text and not text.startswith("#") and not text.startswith("```"):
            current[1].append(("text", text))
        i += 1
    if current[1]:
        sections.append(current)
    return sections


def sheet_name(title, used):
    clean = re.sub(r"[\\/*?:\[\]]", "", title)
    clean = re.sub(r"^(Step\s*[\d.]+:\s*)", "", clean).strip()
    name = clean[:31] or "Sheet"
    n = 2
    while name in used:
        suffix = f"_{n}"
        name = clean[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def write_sheet(ws, title, blocks):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    r = 3
    widths = {}
    for kind, payload in blocks:
        if kind == "subhead":
            c = ws.cell(row=r, column=1, value=payload)
            c.font = Font(bold=True, size=12, color="1F3864")
            r += 2
        elif kind == "text":
            c = ws.cell(row=r, column=1, value=re.sub(r"\*\*", "", payload))
            c.font = NOTE_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
            widths[1] = max(widths.get(1, 0), 40)
            r += 1
        elif kind == "table":
            header, rows = payload
            for ci, h in enumerate(header, start=1):
                c = ws.cell(row=r, column=ci, value=re.sub(r"\*\*", "", h))
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
                c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                c.border = BORDER
                widths[ci] = max(widths.get(ci, 0), min(len(h) + 6, 34))
            r += 1
            for row in rows:
                is_section = all(v in ("", "—", "-") for v in row[1:]) and row[0]
                for ci, v in enumerate(row, start=1):
                    val = re.sub(r"\*\*", "", v).replace("• ", "\n• ").strip()
                    c = ws.cell(row=r, column=ci, value=val)
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                    c.border = BORDER
                    if is_section:
                        c.fill = SECTION_FILL
                        c.font = Font(bold=True)
                    widths[ci] = max(widths.get(ci, 0), min(max(len(x) for x in val.split("\n")) + 2, 46))
                r += 1
            r += 1
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = max(w, 14)
    ws.freeze_panes = "A3"


def main():
    src, dst = sys.argv[1], sys.argv[2]
    overrides = dict(a.split("=", 1) for a in sys.argv[3:])
    sections = parse(open(src).read())
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for title, blocks in sections:
        name = sheet_name(overrides.get(title, title), used)
        write_sheet(wb.create_sheet(name), title, blocks)
    wb.save(dst)
    print(f"{dst}: {len(wb.sheetnames)} sheets -> {wb.sheetnames}")


if __name__ == "__main__":
    main()
